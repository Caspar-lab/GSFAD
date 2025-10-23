import numpy as np
from openpyxl import Workbook
from scipy import io
import time
import warnings
from sklearn.metrics import roc_auc_score
from sklearn.preprocessing import MinMaxScaler
import os
from sklearn.utils import check_consistent_length, column_or_1d
import matplotlib.pyplot as plt
from iNNe_improve._inneGB_10 import IsolationNNE

warnings.filterwarnings('ignore')
# 无二次分裂+去除1的球
if __name__ == '__main__':
    algorithm_name = 'iNNe_GB'
    env = "dev"
    # env = "tes"
    # load the data
    dir_path = os.path.split(os.path.realpath(__file__))[0]
    dataname_path = os.path.join(dir_path, 'datasets\\all_datalists_outlier.mat')  # Categorical Mixed Numerical
    datalists = io.loadmat(dataname_path)['datalists']

    # no_data_ID = [8, 9, 26, 32, 34, 39, 47, 48, 51] + list(range(62, 68))
    no_data_ID = [8, 9, 26, 32, 34, 38, 39, 47, 48, 51] + list(range(62, 68))
    # 生成27-83的所有数字
    # 28从card开始
    # full_range = range(0, 84)  # 注意：range右边界不包含，所以84才能包含83
    # test range
    # 过滤掉no_data中的数字
    # result = [num-1 for num in full_range if num not in no_data_ID]
    # data_ID = [29, 31, 33, 35, 36, 40, 42, 46, 49, 56, 57, 60, 68, 69, 71, 73, 74, 77, 82, 83]
    data_ID = [29, 30, 32, 36, 37, 41, 44, 56, 57, 58, 59, 60, 61, 69, 70, 71, 80, 84]
    data_ID = np.asarray(data_ID) - 1
    avg_auc = []
    selected_Data = ["mushroom_p_221_variant1", "vote_republican_29_variant1", "cardio",
                     "cardiotocography_2and3_33_variant1",
                     "ecoli", "ionosphere_b_24_variant1", "iris_Irisvirginica_11_variant1", "musk", "pendigits",
                     "waveform_0_100_variant1", "wbc_malignant_39_variant1", "wdbc_M_39_variant1", "wine",
                     "wpbc_variant1", "yeast_ERL_5_variant1", "annealing_variant1", "arrhythmia_variant1",
                     "autos_variant1",
                     "hepatitis_2_9_variant1", "thyroid_disease_variant1"]
    for data_i in range(len(datalists)):
        print('data_i=', data_i)
        data_name = datalists[data_i][0][0]
        print('old_dataset:', data_name)
        if data_name in selected_Data:
            if data_i in no_data_ID:
                print('Dataset:' + data_name + ' 运行不出来！！！')
                continue
            add_folder = os.path.join(
                os.path.join(dir_path, 'Exp_results\\' + algorithm_name + '\\results13\\' + data_name))
            if env == "dev":
                if os.path.exists(add_folder):
                    print(data_name + " 已经有实验结果！！！")
                    continue
                os.mkdir(add_folder)

            data_path = os.path.join(dir_path, 'datasets\\' + data_name + '.mat')
            trandata = io.loadmat(data_path)['trandata']

            oridata = trandata.copy()
            trandata = trandata.astype(float)
            # 标准化原始数据
            ID = (trandata >= 1).all(axis=0) & (trandata.max(axis=0) != trandata.min(axis=0))
            scaler = MinMaxScaler()
            if any(ID):
                trandata[:, ID] = scaler.fit_transform(trandata[:, ID])

            X = trandata[:, :-1]  # X是去除标签之后的数据
            print("dataSize:", X.shape)
            labels = trandata[:, -1]

            opt_AUC = 0
            opt_out_scores = np.zeros(len(labels))
            opt_k = 0
            opt_T = 0
            opt_delta = 0

            n_estimators = [
                20
            ]
            # max_samples在[2^2,2^10]之间取值
            # max_samples = [
            #     4, 8, 16, 32, 64, 128
            # ]
            max_samples = [
                4
            ]
            # max_samples = [
            #     256, 512
            # ]
            excel_add = []
            # 循环遍历n_estimators和max_samples的所有组合
            for n in n_estimators:
                for m in max_samples:
                    for k in range(2, 60):
                        # if k >= m:
                        #     continue
                        start_time = time.time()
                        try:
                            clf = IsolationNNE(n_estimators=100, max_samples=m, k=k)
                        except ValueError as e:
                            print(f"Error with n_estimators={n}, max_samples={m}, k={k}: {e}")
                            continue
                        try:
                            clf.fit(X, ID[:-1])
                        except ValueError as e:
                            print(f"Error fitting model with n_estimators={n}, max_samples={m}, k={k}: {e}")
                            continue
                        out_scores, y_pred_train = clf.predict(X)
                        variance = clf.variance_
                        end_time = time.time()
                        execution_time = end_time - start_time
                        results_name1 = 'data_' + str(data_i) + '_' + algorithm_name + '_e' + str(n) + '_s' + str(m) + '_k' + str(k) + '.mat'
                        AUC = roc_auc_score(labels, out_scores)
                        excel_add.append([data_name, n, m, k, AUC, execution_time, variance])
                        save_path = os.path.join(add_folder, results_name1)
                        if env == "dev":
                            io.savemat(save_path, {'out_scores': out_scores})
                        print('n_estimators:', n)
                        print('max_samples:', m)
                        print('k:', k)
                        print('Time:', execution_time)
                        print(f"AUC:{AUC}")
                        if AUC > opt_AUC:
                            opt_k = k
                            opt_AUC = AUC
                            opt_out_scores = out_scores
                            opt_time = execution_time
                            opt_n_estimators = n
                            opt_max_samples = m
                        print(f"opt_AUC:{opt_AUC}")
            # 保存excel_add到excel文件
            # 创建一个新的工作簿
            wb = Workbook()
            ws = wb.active
            for i, value in enumerate(excel_add, start=1):
                # ws[f'A{i}'] = seleted_data_ID[i-1]
                ws[f'A{i}'] = value[0]
                ws[f'B{i}'] = value[1]
                ws[f'C{i}'] = value[2]
                ws[f'D{i}'] = value[3]
                ws[f'E{i}'] = value[4]
                ws[f'F{i}'] = value[5]
                ws[f'G{i}'] = value[6]


            print('opt_AUC=', opt_AUC)
            T_temp = np.zeros((len(opt_out_scores), 1))
            T_temp[0] = opt_AUC
            T_temp[1] = opt_time
            T_temp[2] = opt_n_estimators
            T_temp[3] = opt_max_samples
            T_temp[4] = opt_k
            opt_out_scores = opt_out_scores.reshape(-1, 1)
            # 添加一列实验记录T_temp
            opt_out_scores = np.column_stack((opt_out_scores, T_temp))
            results_name2 = data_name + '_' + algorithm_name + '.mat'
            save_path = os.path.join(add_folder, results_name2)
            if env == "dev":
                io.savemat(save_path, {'opt_out_scores': opt_out_scores})
            avg_auc.append(opt_AUC)
            # 保存文件
            wb.save(add_folder + "\\results.xlsx")
    print('avg_auc=', np.mean(np.asarray(avg_auc)))